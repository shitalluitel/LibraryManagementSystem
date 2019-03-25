import json

from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from django.db import transaction, IntegrityError
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import render, redirect, reverse

# Create your views here.
# from rolepermissions.decorators import has_permission_decorator
from openpyxl import load_workbook

from books.forms import BookForm, BookUnitAddForm, BookUnitDeleteForm
from books.models import Book, BookUnit, BOOK_STATUS
from borrows.models import Borrow
from students.forms import StudentFileForm


def home(request):
    messages.success(request, "Yes")
    return render(request, 'page.html')


@permission_required('books.add_book', raise_exception=True)
def create_book(request):
    context = {}
    form = BookForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            try:
                with transaction.atomic():
                    book = form.save()

                    BookUnit(book=book).save()
                    messages.success(request, "Book {} successfully created.".format(book.name))
                    return redirect('books:book_create')
            except IntegrityError:
                messages.error("Unable to save your request due to some integrity error.")
    context['form'] = form
    return render(request, 'books/create.html', context)


@permission_required('books.add_book', raise_exception=True)
def create_json(request):
    context = {}
    form = BookForm(request.POST or None)

    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    book = form.save()

                    BookUnit(book=book).save()
                    messages.success(request, "Book {} successfully created.".format(book.name))
                    return redirect('books:book_list')
            except IntegrityError:
                messages.error("Unable to save your request due to some integrity error.")
                return create_book(request=request)
        else:
            return create_book(request=request)
    context['form'] = form
    return render(request, 'books/_form.html', context)


@permission_required('books.view_books', login_url='users:login_user')
def list_book(request):
    datas = Book.objects.filter(is_deleted=False)
    url = reverse("books:json_list_book")
    return render(request, 'books/list_book.html', {'url': url, 'status': True, 'datas': datas})


@permission_required('books.view_books', raise_exception=True)
def json_response(request, datas):
    totalRecords = datas.count()
    q_string = request.GET.get('sSearch')
    # for search action in datatables
    # start

    lookup = Q(name__icontains=q_string) | Q(author__icontains=q_string) | Q(
        edition__icontains=q_string) | Q(publisher__icontains=q_string)

    if not q_string is None:
        datas = datas.filter(lookup)

    # ends

    datas = datas.serialize()
    dict_data = json.loads(datas)

    # for pagination
    # start
    # length = int(request.GET.get("length"))
    # start = int(request.GET.get("start"))
    length = int(request.GET.get("iDisplayLength"))
    start = int(request.GET.get("iDisplayStart"))
    dict_data = dict_data[start: start + length]

    context = {'iTotalRecords': totalRecords, 'iTotalDisplayRecords': totalRecords, 'aaData': dict_data}
    json_datas = json.dumps(context)
    return HttpResponse(json_datas, content_type='application/json', status=200)


@permission_required('books.view_books', raise_exception=True)
def json_list(request):
    datas = Book.objects.filter(is_deleted=False)
    return json_response(request=request, datas=datas)


@permission_required('books.delete_book', raise_exception=True)
def book_trash(request):
    url = reverse("books:json_trash_book")
    return render(request, 'books/list_book.html', {'url': url, 'status': False})


@permission_required('books.delete_book', raise_exception=True)
def json_trash(request):
    datas = Book.objects.filter(is_deleted=True)
    return json_response(request=request, datas=datas)


@permission_required('books.change_book', raise_exception=True)
def edit_book(request, pk):
    context = {}
    try:
        data = Book.objects.get(id=pk)
    except Book.DoesNotExist:
        messages.error(request, "")
        return redirect('books:book_list')

    form = BookForm(request.POST or None, instance=data)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Data updated successpully.")
            return redirect("books:book_list")

    context['form'] = form
    return render(request, 'books/create.html', context)


@permission_required('books.delete_book', raise_exception=True)
def delete_book(request, pk):
    context = {}
    if request.method == "POST":
        try:
            data = Book.objects.get(pk=pk)
            book_units = data.book_units.count()
            if book_units == 0:
                data.is_deleted = True
                data.save()
            else:
                messages.error(request,
                               'Unable to delete book. There are {} no. of book units associated with this book.'.format(
                                   book_units))
        except Book.DoesNotExist:
            messages.error(request, "Unable to find data that you have requested.")
            return redirect('books:book_list')

        messages.success(request, "Successfully deleted {}.".format(data.name))
        return redirect('books:book_list')

    context['next'] = reverse('books:book_list')
    context['status'] = "Delete"
    context['page_name'] = "Books"
    return render(request, 'snippets/delete.html', context)


@permission_required('books.undo_book', raise_exception=True)
def undo_book(request, pk):
    context = {}
    if request.method == "POST":
        try:
            data = Book.objects.get(pk=pk)
            data.is_deleted = False
            data.save()
        except Book.DoesNotExist:
            messages.error(request, "Unable to find data that you have requested.")
            return redirect('books:book_list')

        messages.success(request, "Data {} restored successfully.".format(data.name))
        return redirect('books:book_list')

    context['next'] = reverse('books:book_list')
    context['status'] = "Undo"
    context['page_name'] = "Books"
    return render(request, 'snippets/delete.html', context)


@permission_required('books.view_bookunit', raise_exception=True)
def list_book_unit(request, pk):
    try:
        data = Book.objects.get(is_deleted=False, pk=pk)
    except Book.DoesNotExist:
        messages.error(request, 'Unable to find book.')
        return redirect('books:book_list')
    bookunits = data.book_units.filter(is_deleted=False)

    context = {
        'pk': pk,
        'url': reverse("books:list_book_unit_json", kwargs={'pk': pk}),
        'datas': bookunits,
        'create_url': reverse("books:create_book_units", kwargs={'pk': pk})
    }

    return render(request, 'book_units/list_book_unit.html', context)


@permission_required('books.view_bookunit', raise_exception=True)
def list_book_unit_json(request, pk):
    book = None
    try:
        book = Book.objects.get(pk=pk, is_deleted=False)
    except Book.DoesNotExist:
        messages.error(request, "Unable to find book unit for book you have supplied.")
        return redirect("books:list_book")

    book_unit = book.book_units.filter(is_deleted=False).values('acc_no', 'status', 'id')
    context = {'data': list(book_unit)}

    return HttpResponse(json.dumps(context), 'application/json', status=200)


@permission_required('books.add_bookunit', raise_exception=True)
def create_book_units(request, pk):
    context = {}
    form = BookUnitAddForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            book = Book.objects.get(pk=pk)
            units = form.cleaned_data.get('units')

            try:
                with transaction.atomic():
                    for i in range(0, units):
                        BookUnit(book=book).save()
                    return redirect("books:list_book_unit", pk=pk)
            except IntegrityError:
                messages.error("Unable to save your request due to some integrity error.")

    context['form'] = form
    context['pk'] = pk
    return render(request, 'book_units/create.html', context)


@permission_required('books.create_bookunit', raise_exception=True)
def create_book_units_json(request, pk):
    context = {}
    form = BookUnitAddForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            book = Book.objects.get(pk=pk)
            units = form.cleaned_data.get('units')

            try:
                with transaction.atomic():
                    for i in range(0, units):
                        BookUnit(book=book).save()
                    return redirect("books:list_book_unit", pk=pk)
            except IntegrityError:
                messages.error("Unable to save your request due to some integrity error.")
                return create_book_units(request=request, pk=pk)
        else:
            return create_book_units(request=request, pk=pk)

    context['form'] = form
    context['pk'] = pk
    return render(request, 'book_units/_form.html', context)


@permission_required('books.delete_bookunit', raise_exception=True)
def delete_book_units(request, pk):
    context = {}
    form = BookUnitDeleteForm(request.POST or None)
    data = None
    try:
        data = BookUnit.objects.get(pk=pk)
    except BookUnit.DoesNotExist:
        messages.error(request, "Unable to find data that you have requested.")
        return redirect('books:list_book_unit', pk=pk)

    if request.method == "POST":
        if form.is_valid():
            if data.is_available:
                data.remarks = form.cleaned_data.get('remark')
                data.is_deleted = True
                data.save()

                messages.success(request, "Successfully deleted {}.".format(data.acc_no))
                return redirect('books:list_book_unit', pk=data.book.id)
            messages.error(request, "Unable to delete data.")
            return redirect('books:list_book_unit', pk=data.book.id)
        else:
            messages.error(request, "Unable to delete data.")
            return redirect('books:list_book_unit', pk=data.book.id)
    context['form'] = form
    context['next'] = reverse('books:list_book_unit', kwargs={'pk': data.book.id})
    context['status'] = "Delete"
    context['page_name'] = "Book Units"
    return render(request, 'snippets/book_unit_delete.html', context)


@permission_required('books.view_bookunit', raise_exception=True)
def get_bookunit_option(request, pk):
    book_unit = BookUnit.objects.filter(book__id=pk, status='available')
    context = {}
    context['datas'] = book_unit
    context['empty_label'] = 'Book Unit'
    return render(request, 'snippets/book_option.html', context)


@permission_required('books.add_book')
def add_book_from_file(request):
    context = {}
    form = StudentFileForm(data=request.POST or None, files=request.FILES)
    if request.method == 'POST':
        if form.is_valid():
            file = request.FILES.get('document')
            with transaction.atomic():
                file = form.cleaned_data.get('document')

                try:
                    wb = load_workbook(file)
                    sheet = wb['Sheet1']
                    ws = wb.active
                    rows = tuple(ws.rows)
                    for i in range(2, len(rows) + 1):
                        code = sheet.cell(i, 1).value
                        name = sheet.cell(i, 2).value
                        author = sheet.cell(i, 3).value
                        publisher = sheet.cell(i, 4).value
                        edition = sheet.cell(i, 5).value
                        no_of_units = int(sheet.cell(i, 6).value)

                        book = Book.objects.create(code=code, name=name, author=author, publisher=publisher,
                                                   edition=edition)

                        for i in range(0, no_of_units):
                            BookUnit(book=book).save()

                    messages.success(request, "Successfully created books from given file.")
                except Exception as e:
                    print(e)
                    messages.error(request, e)

    context['form'] = form

    return render(request, 'books/create_books_from_file.html', context)


@permission_required('books.add_book')
def book_home(request):
    return render(request, 'books/book_home.html')
