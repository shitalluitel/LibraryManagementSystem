from django.contrib import messages
from django.contrib.auth.decorators import permission_required, login_required

from django.db import IntegrityError, transaction
from django.forms import modelformset_factory
from django.shortcuts import render, redirect, reverse

# Create your views here.
from openpyxl import load_workbook

from settings.models import CourseBatch
from students.forms import CreateChoiceForm, StudentCreateForm, StudentFileForm
from students.models import Student


@permission_required('students.add_student', raise_exception=True)
def create_student(request):
    context = {}

    form = CreateChoiceForm(request.POST or None)
    modelfromset = modelformset_factory(Student, form=StudentCreateForm)
    student_formset = modelfromset(request.POST or None, request.FILES or None, queryset=Student.objects.none(),
                                   prefix='student')
    # form = StudentCreateForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid() and student_formset.is_valid():
            try:
                with transaction.atomic():
                    course = form.cleaned_data.get('course')
                    batch = form.cleaned_data.get('batch')
                    course_batch = CourseBatch.objects.get(course=course, batch=batch)

                    for student_form in student_formset.forms:
                        student = student_form.save(commit=False)
                        student.course_batch = course_batch
                        student.save()

                    return redirect('students:create_student')
            except IntegrityError():
                pass

        else:
            print("{}, {}".format(form.errors, student_formset.errors))

    context['form'] = form
    context['formset'] = student_formset
    return render(request, 'students/create.html', context)


@permission_required('students.view_student', raise_exception=True)
def list_student(request):
    context = {}
    form = CreateChoiceForm(request.POST or None)
    context['form'] = form
    if request.method == 'POST':
        if form.is_valid():
            data = form.cleaned_data
            course = data.get('course')
            batch = data.get('batch')
            student = Student.objects.filter(course_batch__batch=batch, course_batch__course_id=course)
            context['datas'] = student
            return render(request, 'students/list_student.html', context)

    return render(request, 'students/select_course_batch.html', context)


@permission_required('students.view_student', raise_exception=True)
def json_list_student(request):
    data = request.POST
    print(data.get('course'))
    print(data.get('batch'))
    pass


# 
# @permission_required('students.view_student', raise_exception=True)
# def detail_student(request, pk):
#     try:
#         data = Student.objects.get()


@permission_required('students.view_student', raise_exception=True)
def get_student_option(request):
    context = {}
    batch = request.GET.get('batch')
    course = request.GET.get('course')

    student = Student.objects.filter(course_batch__course_id=course, course_batch__batch_id=batch).order_by('name')
    context['students'] = student
    context['empty_label'] = 'Select Student'
    return render(request, 'snippets/student_option.html', context)


@permission_required('students.add_student', raise_exception=True)
def add_student_from_file(request):
    context = {}
    choice_form = CreateChoiceForm(data=request.POST)
    form = StudentFileForm(data=request.POST or None, files=request.FILES)
    if request.method == 'POST':
        if form.is_valid() and choice_form.is_valid():
            file = request.FILES.get('document')
            with transaction.atomic():
                course = choice_form.cleaned_data.get('course')
                batch = choice_form.cleaned_data.get('batch')
                course_batch = CourseBatch.objects.get(course=course, batch=batch)

                file = form.cleaned_data.get('document')

                try:
                    wb = load_workbook(file)
                    sheet = wb['Sheet1']
                    ws = wb.active
                    rows = tuple(ws.rows)
                    for i in range(2, len(rows) + 1):
                        name = sheet.cell(i, 1).value
                        dob = sheet.cell(i, 2).value
                        phone = sheet.cell(i, 3).value
                        address = sheet.cell(i, 4).value

                        student = Student.objects.create(name=name, dob=dob, phone_no=phone, address=address,
                                                         course_batch=course_batch)

                        print("Name: {}, DOB: {}, Phone: {}, Address: {}".format(name, dob, phone, address))
                    # Student.objects.bulk_create(create_list)
                    messages.success(request, "Successfully created student from given file.")
                except Exception as e:
                    print(e)
                    messages.error(request, e)

    context['choice_form'] = choice_form
    context['form'] = form

    return render(request, 'students/student_upload_file.html', context)


@permission_required('students.add_student', raise_exception=True)
def student_home(request):
    return render(request, 'students/student_home.html')
